import tkinter as tk
from tkinter import simpledialog, messagebox
import openpyxl
import os

# Path for the Excel file to store contacts
contacts_file = "TASK 3/CONTACTS FILE/contacts.xlsx"

# Initialize or load the Excel file
def load_or_create_workbook():
    if not os.path.exists(contacts_file):
        os.makedirs(os.path.dirname(contacts_file), exist_ok=True)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Contacts"
        sheet.append(["Sn No.", "Name", "Contact No.", "Email"])
        workbook.save(contacts_file)
    else:
        workbook = openpyxl.load_workbook(contacts_file)
    return workbook

# Save contacts to Excel file
def save_contacts_to_excel():
    workbook = load_or_create_workbook()
    sheet = workbook["Contacts"]

    # Clear the existing rows except the header
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=4):
        for cell in row:
            cell.value = None

    # Write contacts to the Excel file
    for idx, (name, info) in enumerate(contacts.items(), start=1):
        sheet.append([idx, name, info["phone"], info["email"]])

    workbook.save(contacts_file)

# Add a new contact
def add_contact():
    name = simpledialog.askstring("Input", "Enter contact name:", parent=root)
    if name:
        phone = simpledialog.askstring("Input", "Enter contact phone number:", parent=root)
        email = simpledialog.askstring("Input", "Enter contact email address:", parent=root)
        contacts[name] = {"phone": phone, "email": email}
        save_contacts_to_excel()
        refresh_contacts_list()

# Edit an existing contact
def edit_contact():
    selected = contacts_listbox.curselection()
    if selected:
        name = contacts_listbox.get(selected)
        phone = simpledialog.askstring("Input", f"Enter new phone number for {name}:", parent=root)
        email = simpledialog.askstring("Input", f"Enter new email address for {name}:", parent=root)
        contacts[name] = {"phone": phone, "email": email}
        save_contacts_to_excel()
        refresh_contacts_list()

# Delete a contact
def delete_contact():
    selected = contacts_listbox.curselection()
    if selected:
        name = contacts_listbox.get(selected)
        del contacts[name]
        save_contacts_to_excel()
        refresh_contacts_list()

# Refresh the contacts listbox
def refresh_contacts_list():
    contacts_listbox.delete(0, tk.END)
    for name in contacts:
        contacts_listbox.insert(tk.END, name)

# Display contact details
def display_contact(event):
    selected = contacts_listbox.curselection()
    if selected:
        name = contacts_listbox.get(selected)
        contact = contacts[name]
        details = f"Name: {name}\nPhone: {contact['phone']}\nEmail: {contact['email']}"
        messagebox.showinfo("Contact Details", details)

# Load contacts from Excel file
def load_contacts_from_excel():
    workbook = load_or_create_workbook()
    sheet = workbook["Contacts"]
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        if row[1]:  # Check if the name column is not empty
            name = row[1]
            contacts[name] = {"phone": row[2], "email": row[3]}

# Initialize main application window
root = tk.Tk()
root.title("Contact Manager")
root.geometry("800x500")
root.configure(bg="black")

# Initialize contacts dictionary
contacts = {}

# Load contacts
load_contacts_from_excel()

# Create and place widgets
title_label = tk.Label(root, text="Contact Manager", font=("Helvetica", 20), bg="black", fg="white")
title_label.pack(pady=10)

add_button = tk.Button(root, text="Add Contact", command=add_contact, bg="lightgray", font=("Helvetica", 20))
add_button.pack(pady=10)

edit_button = tk.Button(root, text="Edit Contact", command=edit_contact, bg="lightgray", font=("Helvetica", 20))
edit_button.pack(pady=10)

delete_button = tk.Button(root, text="Delete Contact", command=delete_contact, bg="lightgray", font=("Helvetica", 20))
delete_button.pack(pady=10)

contacts_listbox = tk.Listbox(root, font=("Helvetica", 20), bg="black", fg="white")
contacts_listbox.pack(pady=10, fill=tk.BOTH, expand=True)
contacts_listbox.bind('<<ListboxSelect>>', display_contact)

# Populate contacts listbox
refresh_contacts_list()

# Start the Tkinter event loop
root.mainloop()